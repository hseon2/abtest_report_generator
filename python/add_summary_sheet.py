import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def apply_white_background(ws, max_row=200, max_col=30):
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = white_fill


def main():
    if len(sys.argv) < 3:
        print("Usage: python add_summary_sheet.py <excel_path> <summary_json_path>")
        sys.exit(1)

    excel_path = sys.argv[1]
    summary_json_path = sys.argv[2]

    with open(summary_json_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    test_title = (payload.get("testTitle") or "").strip()
    ab_test_summary = (payload.get("abTestSummary") or "").strip()
    ab_test_results = (payload.get("abTestResults") or "").strip()

    wb = load_workbook(excel_path)

    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])

    # 모든 기존 시트 B3에 테스트명 반영
    for sheet_name in wb.sheetnames:
        target_ws = wb[sheet_name]
        target_ws["B3"] = test_title

    ws = wb.create_sheet("Summary", 0)

    apply_white_background(ws, max_row=220, max_col=40)

    # B3: Test Title
    ws["B3"] = test_title
    ws["B3"].font = Font(name="Samsung SS Head KR Bold", size=18, bold=True)

    # B4: 사이트/기간 고정 문구
    ws["B4"] = "- 사이트 / 기간"
    ws["B4"].font = Font(name="Malgun Gothic", size=12, bold=False)

    # 6행: AB 테스트 결과 요약
    ws["B6"] = "AB 테스트 결과 요약"
    ws["B6"].font = Font(name="Malgun Gothic", size=14, bold=True)
    fill_summary_header = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    for col in range(2, 7):  # B~F
        ws.cell(row=6, column=col).fill = fill_summary_header

    # 6행 하위에 요약 항목을 각 행으로 기록
    current_row = 7
    summary_lines = [line.rstrip() for line in ab_test_summary.splitlines() if line.strip()]
    for line in summary_lines:
        ws.cell(row=current_row, column=2, value=line)
        ws.cell(row=current_row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=current_row, column=2).font = Font(name="Malgun Gothic", size=12, bold=False)
        current_row += 1

    # 요약과 결과 사이 1줄 띄움
    if summary_lines:
        current_row += 1

    # AB 테스트 결과 제목 (요약 하위 행 다음)
    result_header_row = current_row
    ws.cell(row=result_header_row, column=2, value="AB 테스트 결과")
    ws.cell(row=result_header_row, column=2).font = Font(name="Malgun Gothic", size=14, bold=True)
    fill_result_header = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    for col in range(2, 7):  # B~F
        ws.cell(row=result_header_row, column=col).fill = fill_result_header
    current_row += 1

    # 결과 항목을 각 행으로 기록
    result_lines = [line.rstrip() for line in ab_test_results.splitlines() if line.strip()]
    for line in result_lines:
        ws.cell(row=current_row, column=2, value=line)
        ws.cell(row=current_row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=current_row, column=2).font = Font(name="Malgun Gothic", size=12, bold=False)
        current_row += 1

    # 가독성 설정
    ws.column_dimensions["B"].width = 120
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[6].height = 24
    for r in range(7, max(current_row + 2, 60)):
        ws.row_dimensions[r].height = 22

    wb.save(excel_path)
    print(f"Summary sheet added to {excel_path}")


if __name__ == "__main__":
    main()

