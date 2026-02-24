#!/usr/bin/env python3
"""
A/B 테스트 리포트 Excel 생성 스크립트
국가별/리포트 순서별 시트 생성 (테스트 정보 영역만)
"""

import sys
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBar, Rule, FormatObject

def create_country_report_order_sheet(wb, country, report_order, country_results, date_range=None, days_live=None, test_title=None):
    """
    국가별/리포트 순서별 시트 생성
    
    Args:
        wb: Workbook 객체
        country: 국가 코드 (CA, UK, CA_FR 등)
        report_order: 리포트 순서 (예: "AUX 2nd report")
        country_results: 해당 국가/리포트 순서의 결과 리스트
        date_range: 날짜 범위 (옵션)
        days_live: 운영 일수 (옵션)
        test_title: 테스트 제목 (옵션)
    """
    # 시트 이름: 국가_리포트순서 (Excel 시트 이름 제한: 31자)
    sheet_name = f"{country}_{report_order}"[:31]
    # Excel 시트 이름에 사용 불가한 문자 제거
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '_')
    
    ws = wb.create_sheet(title=sheet_name)
    
    # 전체 배경 흰색
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # 흰색 테두리
    white_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    # 폰트 정의
    font_body_regular_10 = Font(name="Samsung SS Body KR Regular", size=10, color="000000")
    font_head_bold_18 = Font(name="Samsung SS Head KR Bold", size=18, color="000000", bold=True)
    font_body_regular_9 = Font(name="Samsung SS Body KR Regular", size=9, color="404040")  # 진한 회색
    # 테이블 폰트 (크기 12, Samsung SS Body KR)
    font_table = Font(name="Samsung SS Body KR Regular", size=12, color="000000")
    font_table_bold = Font(name="Samsung SS Body KR Regular", size=12, color="000000", bold=True)
    
    # 행 2: 리포트 순서 (B2) - 앞에 "AUX " 붙이기
    cell = ws['B2']
    if not report_order.startswith("AUX "):
        cell.value = f"AUX {report_order}"
    else:
        cell.value = report_order
    cell.font = font_body_regular_10
    cell.fill = white_fill
    cell.border = white_border
    
    # 행 3: 테스트 명 (B3) - 첫 번째 엑셀 파일명 기준
    cell = ws['B3']
    if test_title:
        cell.value = test_title
    else:
        cell.value = "Test_title"  # test_title이 없으면 "Test_title" 텍스트 표시
    cell.font = font_head_bold_18
    cell.fill = white_fill
    cell.border = white_border
    
    # 행 4: 메타데이터 헤더 (C4:G4)
    ws['C4'] = "Country"
    ws['D4'] = "Date Range"
    ws['E4'] = "Days live"
    ws['F4'] = "Test Segment"
    ws['G4'] = "Daily visit"
    
    for col in ['C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}4']
        cell.font = font_body_regular_9
        cell.fill = white_fill
        cell.border = white_border
    
    # Days 계산: Verdict 계산할 때 분모로 쓰는 days 값 사용
    days_live_value = None
    
    # 디버깅: 입력 파라미터 확인
    print(f"DEBUG create_country_report_order_sheet: country={country}, report_order={report_order}")
    print(f"DEBUG: days_live 파라미터={days_live}, date_range={date_range}")
    print(f"DEBUG: country_results 개수={len(country_results)}")
    
    # country_results에서 days 값 찾기 (우선순위 1)
    for idx, r in enumerate(country_results):
        days_in_result = r.get('days')
        print(f"DEBUG: country_results[{idx}].get('days')={days_in_result}")
        if days_in_result is not None:
            days_live_value = int(days_in_result) if isinstance(days_in_result, (int, float)) else days_in_result
            print(f"DEBUG: days 값 찾음! days_live_value={days_live_value}")
            break
    
    # days_live 파라미터가 있으면 사용 (우선순위 2, country_results에서 찾지 못한 경우만)
    if days_live_value is None and days_live:
        days_live_value = int(days_live) if isinstance(days_live, (int, float)) else days_live
        print(f"DEBUG: days_live 파라미터 사용, days_live_value={days_live_value}")
    
    # days 값을 찾지 못한 경우 date_range에서 계산 (우선순위 3)
    if days_live_value is None and date_range:
        date_range_str = str(date_range).strip()
        
        # "# Date: " 같은 접두사 제거
        if date_range_str.startswith('# Date:'):
            date_range_str = date_range_str.replace('# Date:', '').strip()
        elif date_range_str.startswith('#Date:'):
            date_range_str = date_range_str.replace('#Date:', '').strip()
        elif date_range_str.startswith('Date:'):
            date_range_str = date_range_str.replace('Date:', '').strip()
        
        # 다양한 구분자 시도
        separators = [' - ', ' -', '- ', '-', '~', ' to ', ' To ']
        parts = None
        for sep in separators:
            if sep in date_range_str:
                parts = date_range_str.split(sep, 1)
                break
        
        if not parts and '-' in date_range_str:
            parts = date_range_str.split('-', 1)
        
        if parts and len(parts) == 2:
            try:
                from datetime import datetime as dt
                
                # 여러 날짜 형식 시도
                date_formats = [
                    '%b %d, %Y',     # Apr 16, 2025
                    '%B %d, %Y',     # April 16, 2025
                    '%d %b %Y',      # 16 Apr 2025
                    '%d %B %Y',      # 16 April 2025
                    '%Y-%m-%d',      # 2025-04-16
                    '%Y/%m/%d',      # 2025/04/16
                    '%y/%m/%d',      # 25/04/16
                    '%y-%m-%d',      # 25-04-16
                    '%d/%m/%Y',      # 16/04/2025
                    '%d/%m/%y',      # 16/04/25
                    '%d-%m-%Y',      # 16-04-2025
                    '%d-%m-%y',      # 16-04-25
                    '%m/%d/%Y',      # 04/16/2025
                    '%m/%d/%y',      # 04/16/25
                    '%m-%d-%Y',      # 04-16-2025
                    '%m-%d-%y',      # 04-16-25
                ]
                date1 = None
                date2 = None
                
                # 각 부분 정리 (앞뒤 공백 제거)
                part1 = parts[0].strip()
                part2 = parts[1].strip()
                
                for fmt in date_formats:
                    try:
                        date1 = dt.strptime(part1, fmt)
                        date2 = dt.strptime(part2, fmt)
                        break
                    except:
                        continue
                
                if date1 and date2:
                    days_live_value = (date2 - date1).days + 1
                    print(f"DEBUG: 날짜 계산 성공 - {date1} ~ {date2}, days: {days_live_value}")
                else:
                    print(f"DEBUG: 날짜 파싱 실패 - part1: '{part1}', part2: '{part2}'")
            except Exception as e:
                print(f"날짜 계산 오류: {e}, date_range: {date_range_str}")
                import traceback
                traceback.print_exc()
    
    # 가운데 정렬 스타일
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # 행 5: 메타데이터 값 (C5:G5)
    ws['C5'] = country
    ws['C5'].font = font_body_regular_9
    ws['C5'].fill = white_fill
    ws['C5'].border = white_border
    ws['C5'].alignment = center_alignment
    
    # Date Range (D5)
    print(f"DEBUG: D5 (Date Range) 설정 시작 - date_range={date_range}")
    date_range_value = None
    
    if date_range:
        date_range_value = date_range
        print(f"DEBUG: date_range 파라미터 사용: {date_range_value}")
    else:
        # country_results에서 날짜 정보 찾기 시도
        # 첫 번째 결과에서 날짜 관련 정보 확인
        if country_results and len(country_results) > 0:
            first_result = country_results[0]
            # startDate, endDate 필드 확인
            if first_result.get('startDate') and first_result.get('endDate'):
                start_date = first_result.get('startDate')
                end_date = first_result.get('endDate')
                # YYYY-MM-DD 형식을 YY/MM/DD 형식으로 변환
                try:
                    from datetime import datetime as dt
                    start_dt = dt.strptime(start_date, '%Y-%m-%d')
                    end_dt = dt.strptime(end_date, '%Y-%m-%d')
                    date_range_value = f"{start_dt.strftime('%y/%m/%d')}-{end_dt.strftime('%y/%m/%d')}"
                    print(f"DEBUG: country_results에서 날짜 정보 찾음: {date_range_value}")
                except:
                    pass
    
    if date_range_value:
        ws['D5'] = date_range_value
        print(f"DEBUG: D5 셀에 {date_range_value} 설정됨")
    else:
        # 기본값: 현재 날짜
        default_date = datetime.now().strftime('%y/%m/%d-%y/%m/%d')
        ws['D5'] = default_date
        print(f"DEBUG: D5 셀에 기본값 {default_date} 설정됨 (date_range 없음)")
    
    ws['D5'].font = font_body_regular_9
    ws['D5'].fill = white_fill
    ws['D5'].border = white_border
    ws['D5'].alignment = center_alignment
    
    # Days live (E5) - days_live_value가 None이 아닌 경우에만 설정
    print(f"DEBUG: 최종 days_live_value={days_live_value}")
    if days_live_value is not None:
        ws['E5'] = days_live_value
        print(f"DEBUG: E5 셀에 {days_live_value} 설정됨")
    else:
        ws['E5'] = "N/A"
        print(f"DEBUG: E5 셀에 N/A 설정됨 (days_live_value가 None)")
    ws['E5'].font = font_body_regular_9
    ws['E5'].fill = white_fill
    ws['E5'].border = white_border
    ws['E5'].alignment = center_alignment
    
    # Test Segment (F5) - 비워두기
    ws['F5'] = ""
    ws['F5'].font = font_body_regular_9
    ws['F5'].fill = white_fill
    ws['F5'].border = white_border
    ws['F5'].alignment = center_alignment
    
    # Daily visit 계산은 테이블 생성 후에 수행 (C10, E10, G10 등의 값 사용)
    # 여기서는 일단 "N/A"로 설정하고, 테이블 생성 후에 업데이트
    ws['G5'] = "N/A"
    ws['G5'].font = font_body_regular_9
    ws['G5'].fill = white_fill
    ws['G5'].border = white_border
    ws['G5'].alignment = center_alignment
    
    # 분석 결과 테이블 추가 (7행부터 시작)
    current_row = 7
    
    # 테이블 영역 추적 (테이블이 아닌 영역에 흰색 테두리/배경 적용용)
    table_ranges = []  # [(start_row, end_row, start_col, end_col), ...]
    
    # KPI별로 그룹화
    kpi_groups = {}
    for r in country_results:
        kpi_name = r.get('kpiName', 'Unknown')
        if kpi_name not in kpi_groups:
            kpi_groups[kpi_name] = []
        kpi_groups[kpi_name].append(r)
                
    # KPI 카테고리별 정렬 (primary > secondary > additional)
    def get_category_sort(category):
        sort_order = {'primary': 1, 'secondary': 2, 'additional': 3}
        return sort_order.get(category, 999)
    
    sorted_kpis = sorted(kpi_groups.items(), key=lambda x: get_category_sort(x[1][0].get('category', 'primary')))
    
    # Variation 개수 확인
    variation_count = 1
    for r in country_results:
        if r.get('variations') and len(r.get('variations', [])) > 0:
            variation_count = max(variation_count, len(r.get('variations', [])))
        elif r.get('variationValue') is not None:
            variation_count = 1
    
    print(f"DEBUG: 분석 결과 테이블 생성 시작 - KPI 개수: {len(sorted_kpis)}, Variation 개수: {variation_count}")
    
    # 헤더 생성 함수
    def create_table_header(ws, row, start_col, variation_count, is_variation_only, is_simple_type):
        """테이블 헤더 생성 (2행 구조)"""
        # 첫 번째 헤더 행 (새로운 8행)
        header_row1 = row
        # 두 번째 헤더 행 (새로운 9행)
        header_row2 = row + 1
        
        col = start_col
        
        # 세그먼트 컬럼 (두 행 모두 병합)
        header_cell1 = ws.cell(row=header_row1, column=col)
        header_cell1.value = "세그먼트"
        header_cell1.font = font_table_bold
        header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        header_cell1.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        header_cell1.alignment = center_alignment
        # 세그먼트 열 너비 38로 설정 (라벨 열)
        ws.column_dimensions[get_column_letter(col)].width = 38
        # 세그먼트 셀 병합 (2행)
        ws.merge_cells(f'{get_column_letter(col)}{header_row1}:{get_column_letter(col)}{header_row2}')
        col += 1
        
        if variation_count > 1:
            # 여러 Variation인 경우
            if not is_variation_only:
                # Control - 첫 번째 헤더 행 (C8:D8 병합)
                control_start_col = col
                control_end_col = col + 1
                header_cell1 = ws.cell(row=header_row1, column=control_start_col)
                header_cell1.value = "Control"
                header_cell1.font = font_table_bold
                header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell1.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell1.alignment = center_alignment
                ws.merge_cells(f'{get_column_letter(control_start_col)}{header_row1}:{get_column_letter(control_end_col)}{header_row1}')
                
                # Control Visit - 두 번째 헤더 행 (C9)
                header_cell2 = ws.cell(row=header_row2, column=control_start_col)
                header_cell2.value = "Visit"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                
                # Control % - 두 번째 헤더 행 (D9)
                header_cell2 = ws.cell(row=header_row2, column=control_end_col)
                header_cell2.value = "%"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                col = control_end_col + 1
            
            # Variation 컬럼들
            for i in range(1, variation_count + 1):
                # Variation - 첫 번째 헤더 행 (2열 병합)
                var_start_col = col
                var_end_col = col + 1
                header_cell1 = ws.cell(row=header_row1, column=var_start_col)
                header_cell1.value = f"Variation {i}"
                header_cell1.font = font_table_bold
                header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell1.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell1.alignment = center_alignment
                ws.merge_cells(f'{get_column_letter(var_start_col)}{header_row1}:{get_column_letter(var_end_col)}{header_row1}')
                
                # Variation Visit - 두 번째 헤더 행
                header_cell2 = ws.cell(row=header_row2, column=var_start_col)
                header_cell2.value = "Visit"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                
                # Variation % - 두 번째 헤더 행
                header_cell2 = ws.cell(row=header_row2, column=var_end_col)
                header_cell2.value = "%"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                col = var_end_col + 1
            
            if not is_variation_only:
                # Uplift와 Confidence 컬럼들 (각 Variation마다 2열씩)
                for i in range(1, variation_count + 1):
                    # 첫 번째 헤더 행: "Variation X (Control 대비)" (2열 병합)
                    uplift_start_col = col
                    uplift_end_col = col + 1
                    header_cell1 = ws.cell(row=header_row1, column=uplift_start_col)
                    header_cell1.value = f"Variation {i} (Control 대비)"
                    header_cell1.font = font_table_bold
                    header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                    header_cell1.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    header_cell1.alignment = center_alignment
                    ws.merge_cells(f'{get_column_letter(uplift_start_col)}{header_row1}:{get_column_letter(uplift_end_col)}{header_row1}')
                    
                    # 두 번째 헤더 행: "Uplift" (첫 번째 열)
                    header_cell2 = ws.cell(row=header_row2, column=uplift_start_col)
                    header_cell2.value = "Uplift"
                    header_cell2.font = font_table_bold
                    header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                    header_cell2.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    header_cell2.alignment = center_alignment
                    
                    # 두 번째 헤더 행: "Conf." (두 번째 열)
                    header_cell2 = ws.cell(row=header_row2, column=uplift_end_col)
                    header_cell2.value = "Conf."
                    header_cell2.font = font_table_bold
                    header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                    header_cell2.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    header_cell2.alignment = center_alignment
                    col = uplift_end_col + 1
        else:
            # 단일 Variation인 경우
            if not is_variation_only:
                # Control - 첫 번째 헤더 행 (C8:D8 병합)
                control_start_col = col
                control_end_col = col + 1
                header_cell1 = ws.cell(row=header_row1, column=control_start_col)
                header_cell1.value = "Control"
                header_cell1.font = font_table_bold
                header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell1.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell1.alignment = center_alignment
                ws.merge_cells(f'{get_column_letter(control_start_col)}{header_row1}:{get_column_letter(control_end_col)}{header_row1}')
                
                # Control Visit - 두 번째 헤더 행 (C9)
                header_cell2 = ws.cell(row=header_row2, column=control_start_col)
                header_cell2.value = "Visit"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                
                # Control % - 두 번째 헤더 행 (D9)
                header_cell2 = ws.cell(row=header_row2, column=control_end_col)
                header_cell2.value = "%"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                col = control_end_col + 1
            
            # Variation - 첫 번째 헤더 행 (2열 병합)
            var_start_col = col
            var_end_col = col + 1
            header_cell1 = ws.cell(row=header_row1, column=var_start_col)
            header_cell1.value = "Variation"
            header_cell1.font = font_table_bold
            header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            header_cell1.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            header_cell1.alignment = center_alignment
            ws.merge_cells(f'{get_column_letter(var_start_col)}{header_row1}:{get_column_letter(var_end_col)}{header_row1}')
            
            # Variation Visit - 두 번째 헤더 행
            header_cell2 = ws.cell(row=header_row2, column=var_start_col)
            header_cell2.value = "Visit"
            header_cell2.font = font_table_bold
            header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            header_cell2.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            header_cell2.alignment = center_alignment
            
            # Variation % - 두 번째 헤더 행
            header_cell2 = ws.cell(row=header_row2, column=var_end_col)
            header_cell2.value = "%"
            header_cell2.font = font_table_bold
            header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            header_cell2.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            header_cell2.alignment = center_alignment
            col = var_end_col + 1
            
            if not is_variation_only:
                # Uplift와 Confidence 컬럼 (2열)
                # 첫 번째 헤더 행: "Variation (Control 대비)" (2열 병합)
                uplift_start_col = col
                uplift_end_col = col + 1
                header_cell1 = ws.cell(row=header_row1, column=uplift_start_col)
                header_cell1.value = "Variation (Control 대비)"
                header_cell1.font = font_table_bold
                header_cell1.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell1.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell1.alignment = center_alignment
                ws.merge_cells(f'{get_column_letter(uplift_start_col)}{header_row1}:{get_column_letter(uplift_end_col)}{header_row1}')
                
                # 두 번째 헤더 행: "Uplift" (첫 번째 열)
                header_cell2 = ws.cell(row=header_row2, column=uplift_start_col)
                header_cell2.value = "Uplift"
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                
                # 두 번째 헤더 행: "Conf." (두 번째 열)
                header_cell2 = ws.cell(row=header_row2, column=uplift_end_col)
                header_cell2.value = "Conf."
                header_cell2.font = font_table_bold
                header_cell2.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                header_cell2.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_cell2.alignment = center_alignment
                col = uplift_end_col + 1
        
        return col - start_col  # 컬럼 개수 반환
    
    # 데이터 행 생성 함수 (분모/분자 행 분리)
    def create_data_rows(ws, start_row, start_col, r, variation_count, is_variation_only, is_all_visits, segment_name):
        """데이터 행 생성 (분모 행과 분자 행)"""
        row_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid") if is_all_visits else white_fill
        row_font_weight = 'bold' if is_all_visits else 'normal'
        
        # 분모 행 (첫 번째 행)
        denominator_row = start_row
        data_col = start_col
        
        # 세그먼트 라벨 값 가져오기 (KPI 설정 시 드롭다운으로 선택한 numerator/denominator 라벨)
        # 분모 행에는 denominator 라벨, 분자 행에는 numerator 라벨 사용
        print(f"DEBUG create_data_rows: r.keys()={list(r.keys())}")
        print(f"DEBUG create_data_rows: r.get('numerator')={r.get('numerator')}")
        print(f"DEBUG create_data_rows: r.get('denominator')={r.get('denominator')}")
        print(f"DEBUG create_data_rows: r.get('kpiName')={r.get('kpiName')}")
        
        denominator_label = r.get('denominator', 'N/A')
        if not denominator_label or denominator_label == 'N/A' or denominator_label == '':
            # denominator가 없으면 kpiName을 fallback으로 사용
            denominator_label = r.get('kpiName', 'N/A')
            if not denominator_label or denominator_label == 'N/A':
                denominator_label = r.get('device', 'N/A')
        
        numerator_label = r.get('numerator', 'N/A')
        if not numerator_label or numerator_label == 'N/A' or numerator_label == '':
            # numerator가 없으면 kpiName을 fallback으로 사용
            numerator_label = r.get('kpiName', 'N/A')
            if not numerator_label or numerator_label == 'N/A':
                numerator_label = r.get('device', 'N/A')
        
        print(f"DEBUG create_data_rows: 최종 denominator_label={denominator_label}, numerator_label={numerator_label}")
        
        # 세그먼트 셀 (분모 행) - B열에 denominator 라벨
        segment_cell = ws.cell(row=denominator_row, column=data_col)
        segment_cell.value = denominator_label
        segment_cell.font = font_table_bold
        segment_cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        segment_cell.border = Border(
            left=Side(style='medium', color='3498db'),
            right=Side(style='medium', color='3498db'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        segment_cell.alignment = left_alignment  # 라벨 열은 왼쪽 정렬
        data_col += 1
        
        # 분자 행 (두 번째 행)
        numerator_row = start_row + 1
        num_data_col = start_col
        
        # 세그먼트 셀 (분자 행) - B열에 numerator 라벨
        segment_cell = ws.cell(row=numerator_row, column=num_data_col)
        segment_cell.value = numerator_label
        segment_cell.font = font_table_bold
        segment_cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        segment_cell.border = Border(
            left=Side(style='medium', color='3498db'),
            right=Side(style='medium', color='3498db'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        segment_cell.alignment = left_alignment  # 라벨 열은 왼쪽 정렬
        num_data_col += 1
        
        if variation_count > 1:
            # 여러 Variation인 경우
            variations = r.get('variations', [])
            
            if not is_variation_only:
                # Control 분모 값 (분모 행)
                control_denom = r.get('denominatorSizeControl')
                control_denom_cell = ws.cell(row=denominator_row, column=data_col)
                if control_denom is not None:
                    control_denom_cell.value = int(control_denom)
                else:
                    control_denom_cell.value = "N/A"
                control_denom_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                control_denom_cell.fill = row_fill
                control_denom_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_denom_cell.alignment = center_alignment
                data_col += 1
                
                # Control % 열은 분모 행에 비워둠
                control_pct_denom_cell = ws.cell(row=denominator_row, column=data_col)
                control_pct_denom_cell.value = ""
                control_pct_denom_cell.fill = row_fill
                control_pct_denom_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_pct_denom_cell.alignment = center_alignment
                data_col += 1
                
                # Control 분자 값 (분자 행)
                control_num = r.get('controlValue')
                control_num_cell = ws.cell(row=numerator_row, column=num_data_col)
                if control_num is not None:
                    control_num_cell.value = int(control_num)
                else:
                    control_num_cell.value = "N/A"
                control_num_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                control_num_cell.fill = row_fill
                control_num_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_num_cell.alignment = center_alignment
                num_data_col += 1
                
                # Control % 값 (분모 대비 분자) - 분자 행에 표시
                control_pct_cell = ws.cell(row=numerator_row, column=num_data_col)
                if control_denom is not None and control_num is not None and control_denom > 0:
                    control_pct = (control_num / control_denom) * 100
                    control_pct_cell.value = f"{control_pct:.2f}%"
                else:
                    control_pct_cell.value = "N/A"
                control_pct_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                control_pct_cell.fill = row_fill
                control_pct_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_pct_cell.alignment = center_alignment
                num_data_col += 1
            
            # Variation 값들
            for var_num in range(1, variation_count + 1):
                var_data = next((v for v in variations if v.get('variationNum') == var_num), None)
                
                # Variation 분모 값
                var_denom = var_data.get('denominatorSizeVariation') if var_data else None
                var_denom_cell = ws.cell(row=denominator_row, column=data_col)
                if var_denom is not None:
                    var_denom_cell.value = int(var_denom)
                else:
                    var_denom_cell.value = "N/A"
                    var_denom_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                var_denom_cell.fill = row_fill
                var_denom_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                var_denom_cell.alignment = center_alignment
                data_col += 1
                
                # Variation % 열은 분모 행에 비워둠
                var_pct_denom_cell = ws.cell(row=denominator_row, column=data_col)
                var_pct_denom_cell.value = ""
                var_pct_denom_cell.fill = row_fill
                var_pct_denom_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                var_pct_denom_cell.alignment = center_alignment
                data_col += 1
                
                # Variation 분자 값 (분자 행)
                var_num_val = var_data.get('variationValue') if var_data else None
                var_num_cell = ws.cell(row=numerator_row, column=num_data_col)
                if var_num_val is not None:
                    var_num_cell.value = int(var_num_val)
                else:
                    var_num_cell.value = "N/A"
                    var_num_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                var_num_cell.fill = row_fill
                var_num_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                var_num_cell.alignment = center_alignment
                num_data_col += 1
                
                # Variation % 값 (분모 대비 분자) - 분자 행에 표시
                var_pct_cell = ws.cell(row=numerator_row, column=num_data_col)
                if var_denom is not None and var_num_val is not None and var_denom > 0:
                    var_pct = (var_num_val / var_denom) * 100
                    var_pct_cell.value = f"{var_pct:.2f}%"
                else:
                    var_pct_cell.value = "N/A"
                    var_pct_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                var_pct_cell.fill = row_fill
                var_pct_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                var_pct_cell.alignment = center_alignment
                num_data_col += 1
            
            if not is_variation_only:
                # Uplift와 Confidence 값들 (분자 행에만, 각 Variation마다 2열씩)
                for var_num in range(1, variation_count + 1):
                    var_data = next((v for v in variations if v.get('variationNum') == var_num), None)
                    
                    # Uplift 값 (첫 번째 열)
                    uplift_cell = ws.cell(row=numerator_row, column=num_data_col)
                    if var_data and var_data.get('uplift') is not None:
                        uplift_value = var_data.get('uplift')
                        # 조건부 서식을 위해 숫자 값으로 저장 (퍼센트를 소수로 변환: 5.23% -> 0.0523)
                        uplift_cell.value = uplift_value / 100.0
                        uplift_cell.number_format = '0.00%'  # 퍼센트 형식으로 표시
                        uplift_cell.font = Font(name="Samsung SS Body KR Regular", size=12, 
                                              color="e74c3c" if uplift_value < 0 else "233ffa", 
                                              bold=row_font_weight == 'bold')
                    else:
                        uplift_cell.value = None  # 조건부 서식을 위해 None 또는 0으로 설정
                        uplift_cell.font = Font(name="Samsung SS Body KR Regular", size=12, color="999999", bold=row_font_weight == 'bold')
                    uplift_cell.fill = row_fill
                    uplift_cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    uplift_cell.alignment = right_alignment  # Uplift 셀은 오른쪽 정렬
                    num_data_col += 1
                    
                    # Confidence 값 (두 번째 열)
                    conf_cell = ws.cell(row=numerator_row, column=num_data_col)
                    if var_data and var_data.get('confidence') is not None:
                        confidence = var_data.get('confidence')
                        conf_cell.value = f"{confidence:.2f}%"
                        conf_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                    else:
                        conf_cell.value = "N/A"
                        conf_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                    conf_cell.fill = row_fill
                    conf_cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    conf_cell.alignment = center_alignment
                    num_data_col += 1
        else:
            # 단일 Variation인 경우
            if not is_variation_only:
                # Control 분모 값 (분모 행)
                control_denom = r.get('denominatorSizeControl')
                control_denom_cell = ws.cell(row=denominator_row, column=data_col)
                if control_denom is not None:
                    control_denom_cell.value = int(control_denom)
                else:
                    control_denom_cell.value = "N/A"
                control_denom_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                control_denom_cell.fill = row_fill
                control_denom_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_denom_cell.alignment = center_alignment
                data_col += 1
                
                # Control % 열은 분모 행에 비워둠
                control_pct_denom_cell = ws.cell(row=denominator_row, column=data_col)
                control_pct_denom_cell.value = ""
                control_pct_denom_cell.fill = row_fill
                control_pct_denom_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_pct_denom_cell.alignment = center_alignment
                data_col += 1
                
                # Control 분자 값 (분자 행)
                control_num = r.get('controlValue')
                control_num_cell = ws.cell(row=numerator_row, column=num_data_col)
                if control_num is not None:
                    control_num_cell.value = int(control_num)
                else:
                    control_num_cell.value = "N/A"
                control_num_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                control_num_cell.fill = row_fill
                control_num_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_num_cell.alignment = center_alignment
                num_data_col += 1
                
                # Control % 값 (분모 대비 분자) - 분자 행에 표시
                control_pct_cell = ws.cell(row=numerator_row, column=num_data_col)
                if control_denom is not None and control_num is not None and control_denom > 0:
                    control_pct = (control_num / control_denom) * 100
                    control_pct_cell.value = f"{control_pct:.2f}%"
                else:
                    control_pct_cell.value = "N/A"
                control_pct_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                control_pct_cell.fill = row_fill
                control_pct_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                control_pct_cell.alignment = center_alignment
                num_data_col += 1
            
            # Variation 분모 값 (분모 행)
            var_denom = r.get('denominatorSizeVariation')
            var_denom_cell = ws.cell(row=denominator_row, column=data_col)
            if var_denom is not None:
                var_denom_cell.value = int(var_denom)
            else:
                var_denom_cell.value = "N/A"
            var_denom_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
            var_denom_cell.fill = row_fill
            var_denom_cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            var_denom_cell.alignment = center_alignment
            data_col += 1
            
            # Variation % 열은 분모 행에 비워둠
            var_pct_denom_cell = ws.cell(row=denominator_row, column=data_col)
            var_pct_denom_cell.value = ""
            var_pct_denom_cell.fill = row_fill
            var_pct_denom_cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            var_pct_denom_cell.alignment = center_alignment
            data_col += 1
            
            # Variation 분자 값 (분자 행)
            var_num_val = r.get('variationValue')
            var_num_cell = ws.cell(row=numerator_row, column=num_data_col)
            if var_num_val is not None:
                var_num_cell.value = int(var_num_val)
            else:
                var_num_cell.value = "N/A"
            var_num_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
            var_num_cell.fill = row_fill
            var_num_cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            var_num_cell.alignment = center_alignment
            num_data_col += 1
            
            # Variation % 값 (분모 대비 분자) - 분자 행에 표시
            var_pct_cell = ws.cell(row=numerator_row, column=num_data_col)
            if var_denom is not None and var_num_val is not None and var_denom > 0:
                var_pct = (var_num_val / var_denom) * 100
                var_pct_cell.value = f"{var_pct:.2f}%"
            else:
                var_pct_cell.value = "N/A"
            var_pct_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
            var_pct_cell.fill = row_fill
            var_pct_cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            var_pct_cell.alignment = center_alignment
            num_data_col += 1
            
            if not is_variation_only:
                # Uplift 값 (분자 행에만, 첫 번째 열)
                uplift_cell = ws.cell(row=numerator_row, column=num_data_col)
                if r.get('uplift') is not None:
                    uplift_value = r.get('uplift')
                    # 조건부 서식을 위해 숫자 값으로 저장 (퍼센트를 소수로 변환: 5.23% -> 0.0523)
                    uplift_cell.value = uplift_value / 100.0
                    uplift_cell.number_format = '0.00%'  # 퍼센트 형식으로 표시
                    uplift_cell.font = Font(name="Samsung SS Body KR Regular", size=12, 
                                          color="e74c3c" if uplift_value < 0 else "233ffa", 
                                          bold=row_font_weight == 'bold')
                else:
                    uplift_cell.value = None  # 조건부 서식을 위해 None 또는 0으로 설정
                    uplift_cell.font = Font(name="Samsung SS Body KR Regular", size=12, color="999999", bold=row_font_weight == 'bold')
                uplift_cell.fill = row_fill
                uplift_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                uplift_cell.alignment = right_alignment  # Uplift 셀은 오른쪽 정렬
                num_data_col += 1
                
                # Confidence 값 (분자 행에만, 두 번째 열)
                conf_cell = ws.cell(row=numerator_row, column=num_data_col)
                if r.get('confidence') is not None:
                    confidence = r.get('confidence')
                    conf_cell.value = f"{confidence:.2f}%"
                    conf_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                else:
                    conf_cell.value = "N/A"
                    conf_cell.font = font_table_bold if row_font_weight == 'bold' else font_table
                conf_cell.fill = row_fill
                conf_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                conf_cell.alignment = center_alignment
                num_data_col += 1
        
        return max(data_col, num_data_col) - start_col  # 컬럼 개수 반환
    
    # 각 KPI별로 테이블 생성
    for kpi_name, kpi_results in sorted_kpis:
        first_result = kpi_results[0]
        kpi_category = first_result.get('category', 'primary')
        category_label = {'primary': '[Primary KPI]', 'secondary': '[Secondary KPI]', 'additional': '[Additional Data]'}.get(kpi_category, '')
        
        is_variation_only = (first_result.get('controlValue') is None or first_result.get('controlValue') == '') and \
                           (first_result.get('variationValue') is not None or (first_result.get('variations') and len(first_result.get('variations', [])) > 0))
        is_simple_type = first_result.get('controlRate') is None and first_result.get('controlValue') is not None
        
        # KPI 제목 행 (B7)
        title_cell = ws.cell(row=current_row, column=2)
        title_cell.value = f"{category_label} {kpi_name}" if category_label else kpi_name
        # 폰트: Samsung SS Head KR Bold, 사이즈 12, RGB(68, 114, 196)
        title_cell.font = Font(name="Samsung SS Head KR Bold", size=12, color="4472C4", bold=True)  # RGB(68, 114, 196) = #4472C4
        title_cell.fill = white_fill
        title_cell.border = white_border
        current_row += 1
        
        # 세그먼트별로 그룹화
        segment_groups = {}
        for r in kpi_results:
            # 에러 메시지가 있는 경우는 별도 처리
            if r.get('error') and r.get('errorMessage'):
                if 'error' not in segment_groups:
                    segment_groups['error'] = []
                segment_groups['error'].append(r)
            else:
                device = r.get('device', 'All')
                if device not in segment_groups:
                    segment_groups[device] = []
                segment_groups[device].append(r)
        
        # All visits를 첫 번째로, 나머지는 정렬
        sorted_segments = []
        if 'All' in segment_groups:
            sorted_segments.append(('All', segment_groups['All']))
        for device in sorted(segment_groups.keys()):
            if device != 'All' and device != 'error':
                sorted_segments.append((device, segment_groups[device]))
        if 'error' in segment_groups:
            sorted_segments.append(('error', segment_groups['error']))
        
        # 첫 번째 세그먼트 테이블 (왼쪽)
        data_end_row = current_row  # 초기화
        data_start_row = current_row  # 초기화
        if len(sorted_segments) > 0:
            first_segment_name, first_segment_results = sorted_segments[0]
            start_col = 2  # B열부터 시작
            header_row = current_row
            
            # 에러 메시지 처리
            if first_segment_name == 'error':
                error_cell = ws.cell(row=header_row, column=start_col)
                error_cell.value = f"⚠️ {first_segment_results[0].get('errorMessage')}"
                error_cell.font = font_table_bold
                error_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                error_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                error_cell.alignment = center_alignment
                # 에러 메시지가 여러 컬럼에 걸치도록 병합 (대략적인 너비)
                estimated_cols = 10
                ws.merge_cells(f'B{header_row}:{get_column_letter(start_col + estimated_cols - 1)}{header_row}')
                current_row += 1
                data_start_row = header_row
                data_end_row = current_row - 1
            else:
                # 헤더 생성 (2행 구조)
                num_cols = create_table_header(ws, header_row, start_col, variation_count, is_variation_only, is_simple_type)
                table_start_col = start_col
                table_end_col = start_col + num_cols - 1
                current_row += 2  # 헤더가 2행이므로 2 증가
                
                # 데이터 행들 (분모/분자 행으로 분리)
                data_start_row = current_row
                for idx, r in enumerate(first_segment_results):
                    is_all_visits = (first_segment_name == 'All' or first_segment_name.lower().find('all visits') >= 0) and idx == 0
                    create_data_rows(ws, current_row, start_col, r, variation_count, is_variation_only, is_all_visits, first_segment_name)
                    current_row += 2  # 분모 행과 분자 행 두 행 사용
                data_end_row = current_row - 1
                
                # 테이블 영역 기록
                table_ranges.append((header_row, data_end_row, table_start_col, table_end_col))
                
                # 다음 세그먼트들을 오른쪽에 배치
                if len(sorted_segments) > 1:
                    next_col = table_end_col + 2  # 한 열 간격
                    
                    for seg_idx, (segment_name, segment_results) in enumerate(sorted_segments[1:], 1):
                        # 에러 메시지 처리
                        if segment_name == 'error':
                            error_cell = ws.cell(row=header_row, column=next_col)
                            error_cell.value = f"⚠️ {segment_results[0].get('errorMessage')}"
                            error_cell.font = font_table_bold
                            error_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            error_cell.border = Border(
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                            error_cell.alignment = center_alignment
                            estimated_cols = 10
                            ws.merge_cells(f'{get_column_letter(next_col)}{header_row}:{get_column_letter(next_col + estimated_cols - 1)}{header_row}')
                            next_col += estimated_cols + 1
                        else:
                            # 헤더 생성
                            seg_header_row = header_row
                            seg_num_cols = create_table_header(ws, seg_header_row, next_col, variation_count, is_variation_only, is_simple_type)
                            seg_table_start_col = next_col
                            seg_table_end_col = next_col + seg_num_cols - 1
                            
                            # 데이터 행들 (첫 번째 세그먼트와 같은 행에 맞춤, 분모/분자 행으로 분리)
                            seg_data_row = data_start_row
                            for idx, r in enumerate(segment_results):
                                is_all_visits = (segment_name == 'All' or segment_name.lower().find('all visits') >= 0) and idx == 0
                                if seg_data_row <= data_end_row:
                                    create_data_rows(ws, seg_data_row, next_col, r, variation_count, is_variation_only, is_all_visits, segment_name)
                                    seg_data_row += 2  # 분모 행과 분자 행 두 행 사용
                                else:
                                    # 첫 번째 세그먼트보다 행이 많으면 추가
                                    create_data_rows(ws, seg_data_row, next_col, r, variation_count, is_variation_only, is_all_visits, segment_name)
                                    seg_data_row += 2  # 분모 행과 분자 행 두 행 사용
                                    data_end_row = seg_data_row - 1
                            
                            # 테이블 영역 기록
                            table_ranges.append((seg_header_row, data_end_row, seg_table_start_col, seg_table_end_col))
                            
                            next_col = seg_table_end_col + 2  # 한 열 간격
        
        # KPI 사이 간격
        current_row = max(current_row, data_end_row + 2) if len(sorted_segments) > 0 else current_row + 1
    
    # Daily visit 계산 (테이블 생성 후 C10, E10, G10 등의 실제 셀 값 사용)
    # 첫 번째 KPI의 첫 번째 세그먼트 테이블에서 분모 행(10행)의 셀 값을 읽어서 계산
    print(f"DEBUG: Daily visit 계산 시작 (테이블 생성 후 셀 값 읽기)...")
    daily_visit_value = None
    
    # 첫 번째 KPI의 첫 번째 세그먼트 테이블 찾기
    if len(sorted_kpis) > 0:
        first_kpi_name, first_kpi_results = sorted_kpis[0]
        # "Visits" 관련 KPI 찾기
        for r in first_kpi_results:
            kpi_name = str(r.get('kpiName', '')).lower()
            device = r.get('device', 'All')
            
            if 'visit' in kpi_name and (device == 'All' or device == '' or device is None):
                # 테이블의 분모 행(10행)에서 실제 셀 값 읽기
                # C10: Control Visit (분모), E10: Variation 1 Visit (분모), G10: Variation 2 Visit (분모)...
                total_visits = 0
                
                # Control Visit (C10) - 분모 행
                try:
                    c10_value = ws['C10'].value
                    if c10_value is not None:
                        c10_num = float(c10_value) if isinstance(c10_value, (int, float, str)) and str(c10_value).replace(',', '').replace('.', '').isdigit() else 0
                        total_visits += c10_num
                        print(f"DEBUG: C10 (Control) 값: {c10_value} -> {c10_num}, 총합: {total_visits}")
                except:
                    print(f"DEBUG: C10 셀 읽기 실패")
                
                # Variation Visits
                # Variation이 1개인 경우: E10
                # Variation이 2개 이상인 경우: E10, G10, ...
                variation_cols = ['E', 'G', 'I', 'K', 'M', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG', 'AI', 'AK', 'AM', 'AO', 'AQ', 'AS']
                
                if variation_count == 1:
                    # Variation이 1개인 경우: E10
                    try:
                        e10_value = ws['E10'].value
                        if e10_value is not None:
                            e10_num = float(e10_value) if isinstance(e10_value, (int, float, str)) and str(e10_value).replace(',', '').replace('.', '').isdigit() else 0
                            total_visits += e10_num
                            print(f"DEBUG: E10 (Variation 1) 값: {e10_value} -> {e10_num}, 총합: {total_visits}")
                    except:
                        print(f"DEBUG: E10 셀 읽기 실패")
                else:
                    # Variation이 2개 이상인 경우: E10, G10, ...
                    for i in range(variation_count):
                        if i < len(variation_cols):
                            col_letter = variation_cols[i]
                            cell_address = f"{col_letter}10"
                            try:
                                cell_value = ws[cell_address].value
                                if cell_value is not None:
                                    cell_num = float(cell_value) if isinstance(cell_value, (int, float, str)) and str(cell_value).replace(',', '').replace('.', '').isdigit() else 0
                                    total_visits += cell_num
                                    print(f"DEBUG: {cell_address} (Variation {i+1}) 값: {cell_value} -> {cell_num}, 총합: {total_visits}")
                            except:
                                print(f"DEBUG: {cell_address} 셀 읽기 실패")
                
                # Daily visit = 전체 Visit 합산 / 테스트 일수 (소숫점 첫째 자리에서 반올림)
                if days_live_value and isinstance(days_live_value, (int, float)) and days_live_value > 0 and total_visits > 0:
                    daily_visit_value = round(total_visits / days_live_value)
                    print(f"DEBUG: Daily visit 계산 완료: {total_visits} / {days_live_value} = {daily_visit_value}")
                else:
                    print(f"DEBUG: Daily visit 계산 실패 (total_visits={total_visits}, days_live_value={days_live_value})")
                break
    
    # G5 셀에 Daily visit 값 설정
    if daily_visit_value is not None:
        ws['G5'] = daily_visit_value
        ws['G5'].number_format = FORMAT_NUMBER_COMMA_SEPARATED1  # 천단위 구분
        print(f"DEBUG: G5 셀에 {daily_visit_value} 설정됨")
    else:
        print(f"DEBUG: G5 셀에 N/A 유지 (daily_visit_value={daily_visit_value})")
    
    # A1:AT1000 범위의 모든 셀에 흰색 테두리 및 배경 적용 (테이블 셀은 제외)
    def is_in_table(row, col):
        """셀이 테이블 영역에 속하는지 확인"""
        for table_start_row, table_end_row, table_start_col, table_end_col in table_ranges:
            if table_start_row <= row <= table_end_row and table_start_col <= col <= table_end_col:
                return True
        return False
    
    for row in range(1, 1001):
        for col in range(1, 47):  # A=1, AT=46
            cell = ws.cell(row=row, column=col)
            # 테이블 영역이 아닌 경우에만 흰색 테두리/배경 적용
            if not is_in_table(row, col):
                # 이미 스타일이 적용된 셀은 제외 (테이블 셀, 제목 셀 등)
                if cell.border.left.style is None or cell.border.left.color == '00000000':
                    cell.border = white_border
                if cell.fill.start_color.rgb == '00000000' or cell.fill.start_color.rgb == 'FFFFFFFF':
                    cell.fill = white_fill
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 7
    
    # 라벨 열(B, D, N 등) 너비 38로 설정
    # 각 세그먼트 테이블의 시작 열(B열)을 38로 설정
    label_cols = set()  # 라벨 열 추적
    for table_start_row, table_end_row, table_start_col, table_end_col in table_ranges:
        # 각 테이블의 시작 열(B열)이 라벨 열
        label_cols.add(table_start_col)
    
    # 라벨 열 너비 38로 설정
    for label_col in label_cols:
        col_letter = get_column_letter(label_col)
        ws.column_dimensions[col_letter].width = 38
    
    # 나머지 열 너비 19로 설정
    for col in range(1, 47):  # A=1, AT=46
        if col not in label_cols:  # 라벨 열이 아닌 경우
            col_letter = get_column_letter(col)
            if col_letter not in ws.column_dimensions or ws.column_dimensions[col_letter].width != 38:
                ws.column_dimensions[col_letter].width = 19
    
    # Uplift 셀에 조건부 서식 적용
    # 모든 Uplift 셀 범위를 찾아서 조건부 서식 적용
    # Uplift 셀은 각 Variation마다 첫 번째 열에 위치
    # 테이블 범위를 기반으로 Uplift 셀 범위 계산
    uplift_ranges = []  # Uplift 셀 범위 저장 (중복 방지를 위해 set 사용)
    uplift_cols_found = set()  # 이미 찾은 (table_start_row, col) 조합 저장
    for table_start_row, table_end_row, table_start_col, table_end_col in table_ranges:
        # 테이블 내에서 Uplift 셀 위치 찾기 (헤더에서 "Uplift" 찾기)
        # 헤더 행은 table_start_row, table_start_row+1
        # 데이터 행은 table_start_row+2부터 시작 (분모 행)
        # 분자 행은 table_start_row+3부터 시작 (홀수 행)
        # Uplift 컬럼은 헤더에서 "Uplift" 텍스트로 찾기
        for header_row in [table_start_row, table_start_row + 1]:
            for col in range(table_start_col, table_end_col + 1):
                # 이미 찾은 열인지 확인 (같은 테이블 내에서 중복 방지)
                if (table_start_row, col) in uplift_cols_found:
                    continue
                    
                cell = ws.cell(row=header_row, column=col)
                if cell.value and 'Uplift' in str(cell.value):
                    # 이 열이 Uplift 열임
                    # 데이터 행 범위 찾기 (분자 행만: table_start_row+3부터 table_end_row까지, 2행 간격)
                    data_start = table_start_row + 2  # 첫 번째 분모 행
                    # 분자 행만 선택 (홀수 행: data_start+1, data_start+3, ...)
                    numerator_rows = []
                    for row in range(data_start + 1, table_end_row + 1, 2):  # 2행 간격으로 분자 행만
                        numerator_rows.append(row)
                    
                    if numerator_rows:
                        # 첫 번째와 마지막 분자 행으로 범위 생성
                        first_numerator_row = numerator_rows[0]
                        last_numerator_row = numerator_rows[-1]
                        uplift_ranges.append((first_numerator_row, last_numerator_row, col))
                        uplift_cols_found.add((table_start_row, col))  # 찾은 열 기록
                        print(f"DEBUG: Uplift 열 발견 - 테이블 시작 행: {table_start_row}, 열: {get_column_letter(col)}, 분자 행 범위: {first_numerator_row}~{last_numerator_row}")
    
    # Uplift 셀에 조건부 서식 적용
    for data_start_row, data_end_row, uplift_col in uplift_ranges:
        # Uplift 셀 범위 (분자 행만)
        uplift_range = f"{get_column_letter(uplift_col)}{data_start_row}:{get_column_letter(uplift_col)}{data_end_row}"

        # Uplift 셀은 이미 숫자 값으로 저장되어 있으므로 변환 불필요

        # 데이터 막대 조건부 서식 생성
        # 최소값 -0.5, 최대값 0.5 (즉, -50% ~ 50%)
        cfvo = [
            FormatObject(type="num", val=-0.5),
            FormatObject(type="num", val=0.5),
        ]

        # DataBar 기본 설정 (생성자에는 안전한 인자만 사용)
        data_bar = DataBar(
            cfvo=cfvo,
            color="0070C0",  # 양수 막대 색: RGB(0,112,192)
            showValue=True,
            minLength=0,
            maxLength=100,
        )

        # 가능한 경우에만 추가 속성 적용 (openpyxl 버전별 호환성 고려)
        # 음수 막대 색 / 테두리 색
        if hasattr(data_bar, "negativeBarColorSameAsPositive"):
            data_bar.negativeBarColorSameAsPositive = False
        if hasattr(data_bar, "negativeBarBorderColorSameAsPositive"):
            data_bar.negativeBarBorderColorSameAsPositive = False
        if hasattr(data_bar, "negativeBarColor"):
            data_bar.negativeBarColor = "C00000"  # RGB(192,0,0)
        if hasattr(data_bar, "negativeBarBorderColor"):
            data_bar.negativeBarBorderColor = "C00000"  # RGB(192,0,0)

        # 테두리
        if hasattr(data_bar, "border"):
            data_bar.border = True

        # 축 설정: 셀 중간, 축 색상
        if hasattr(data_bar, "axisPosition"):
            data_bar.axisPosition = "cellMid"
        if hasattr(data_bar, "axisColor"):
            data_bar.axisColor = "A5A5A5"  # RGB(165,165,165)

        rule = Rule(type="dataBar", dataBar=data_bar)

        # 조건부 서식 적용
        ws.conditional_formatting.add(uplift_range, rule)
        print(f"DEBUG: Uplift 조건부 서식 적용 - 범위: {uplift_range}")

def create_excel_report(results_path):
    """Excel 리포트 생성 (국가별/리포트 순서별 시트 분리)"""
    print(f"\n{'='*60}")
    print(f"DEBUG: Excel 리포트 생성 시작")
    print(f"DEBUG: results_path={results_path}")
    
    # 결과 로드
    print(f"DEBUG: results.json 파일 읽기 시작...")
    with open(results_path, 'r', encoding='utf-8') as f:
        results = json.load(f)
    print(f"DEBUG: results.json 파일 읽기 완료")
    
    # 결과 구조 확인
    print(f"DEBUG: results 키 목록: {list(results.keys())}")
    if results.get('primaryResults'):
        print(f"DEBUG: primaryResults 개수: {len(results['primaryResults'])}")
        # days 값이 포함된 결과 확인
        results_with_days = [r for r in results['primaryResults'] if r.get('days') is not None]
        print(f"DEBUG: days 값이 포함된 결과: {len(results_with_days)}/{len(results['primaryResults'])}개")
        if results_with_days:
            print(f"DEBUG: 첫 번째 결과의 days 값: {results_with_days[0].get('days')}")
    else:
        print(f"DEBUG: 경고 - primaryResults가 없습니다!")
    
    # Excel 워크북 생성
    print(f"DEBUG: Excel 워크북 생성 시작...")
    wb = Workbook()
    # 기본 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    print(f"DEBUG: Excel 워크북 생성 완료")
    
    # 날짜 범위 추출 (결과에서)
    print(f"DEBUG: 메타데이터 추출 시작...")
    date_range = None
    days_live = None
    test_title = None
    if results.get('metadata'):
        date_range = results['metadata'].get('dateRange')
        days_live = results['metadata'].get('daysLive')
        test_title = results['metadata'].get('testTitle')
        print(f"DEBUG: date_range={date_range}, days_live={days_live}, test_title={test_title}")
    else:
        print(f"DEBUG: 경고 - metadata가 없습니다!")
    
    # Primary Results가 있으면 국가별/리포트 순서별로 그룹화
    if results.get('primaryResults') and len(results['primaryResults']) > 0:
        print(f"DEBUG: Primary Results 그룹화 시작...")
        # 리포트 순서별로 그룹화
        report_order_groups = {}
        for r in results['primaryResults']:
            report_order = r.get('reportOrder', '1st report')
            if report_order not in report_order_groups:
                report_order_groups[report_order] = []
            report_order_groups[report_order].append(r)
        
        print(f"DEBUG: 발견된 리포트 순서: {list(report_order_groups.keys())}")
        print(f"DEBUG: 리포트 순서별 결과 개수:")
        for ro, res in report_order_groups.items():
            print(f"DEBUG:   {ro}: {len(res)}개")
        
        # 리포트 순서 정렬
        sorted_report_orders = sorted(report_order_groups.keys(), key=lambda x: (
            int(x.split()[0]) if x.split()[0].isdigit() else 999
        ))
        print(f"DEBUG: 정렬된 리포트 순서: {sorted_report_orders}")
        
        # 각 리포트 순서별로 처리
        for report_order in sorted_report_orders:
            print(f"\nDEBUG: 리포트 순서 '{report_order}' 처리 시작...")
            report_results = report_order_groups[report_order]
            
            # 국가별로 그룹화
            country_groups = {}
            for r in report_results:
                country = r.get('country', 'Unknown')
                if country not in country_groups:
                    country_groups[country] = []
                country_groups[country].append(r)
            
            sorted_countries = sorted(country_groups.keys())
            print(f"DEBUG: 발견된 국가: {sorted_countries}")
            print(f"DEBUG: 국가별 결과 개수:")
            for country, res in country_groups.items():
                print(f"DEBUG:   {country}: {len(res)}개")
            
            # 각 국가별로 시트 생성
            for country in sorted_countries:
                country_results = country_groups[country]
                print(f"DEBUG: 시트 생성 시작 - 국가: {country}, 리포트 순서: {report_order}, 결과 개수: {len(country_results)}")
                create_country_report_order_sheet(
                    wb, country, report_order, country_results, 
                    date_range, days_live, test_title
                )
                print(f"DEBUG: 시트 생성 완료 - {country}_{report_order}")
    else:
        print(f"DEBUG: 경고 - Primary Results가 없거나 비어있습니다!")
    
    # 파일 저장
    print(f"\nDEBUG: Excel 파일 저장 시작...")
    tmp_dir = Path(results_path).parent
    excel_path = tmp_dir / 'report.xlsx'
    print(f"DEBUG: 저장 경로: {excel_path}")
    wb.save(excel_path)
    print(f"DEBUG: Excel 파일 저장 완료")
    
    # 생성된 시트 확인
    print(f"DEBUG: 생성된 시트 목록: {wb.sheetnames}")
    print(f"DEBUG: 총 시트 개수: {len(wb.sheetnames)}")
    
    print(f"Excel report saved to {excel_path}")
    print(f"{'='*60}\n")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python report_excel.py <results_json>")
        sys.exit(1)
    
    create_excel_report(sys.argv[1])
